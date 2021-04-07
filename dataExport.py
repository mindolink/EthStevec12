from openpyxl import load_workbook


class ExportDada():

    def exportData (self,P,Pstb,Wstb,Pcar,Wcar):
        wb = load_workbook('test.xlsx')
        ws = wb['Sheet1']
        ws['A1'] = 'A1'
        wb.save('test.xlsx')s