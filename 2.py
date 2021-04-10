import openpyxl
from openpyxl import Workbook, worksheet, load_workbook
from openpyxl.utils import get_column_letter

class exportMeausurment(object):

    def __init__(self, userNumber,testNumber,numberOfCars):
        self.numberOfCar=numberOfCars
        self.x=3
        self.filePathName='./ExportData/Data User'+str(userNumber)+' Test'+str(testNumber)+'.xlsx'
        wb = Workbook()
        worksheet = wb.active

        worksheet["B"+ str(3)] = "Day"
        worksheet["C"+ str(3)] = "Haur" 
        worksheet["D"+ str(3)] = "Min"
        worksheet["E"+ str(3)] = "MonayWallet[EURO]"
        worksheet["F"+ str(3)] = "Psum [kW]" 
        worksheet["G"+ str(3)] = "Pl [kW]"
        worksheet["H"+ str(3)] = "Pg [kW]"
        worksheet["I"+ str(3)] = "Esum [kW]" 
        worksheet["J"+ str(3)] = "El [kW]"
        worksheet["J"+ str(3)] = "Eg [kW]"
        worksheet["K"+ str(3)] = "Phsb [kW]"
        worksheet["L"+ str(3)] = "Eg [kW]"      
        worksheet["M"+ str(3)] = "SOChsb [%]"

        for q in range (self.numberOfCar):
            colume= get_column_letter(12+2*q)
            worksheet[str(colume)+str(3)] = "Pcar"+str(q+1)+" [kW]"
            worksheet[str(colume)+str(3) = "Ecar"+str(q+1)+" [kW]"           
            worksheet[str(colume)+str(3)] = "SOCcar"+str(q+1)+" [kW]"   

        wb.save(filename = self.filePathName)

    def homeMeasurment(self,MonayWallet,Psum,Pl,Pg):

        self.x+=1
        wb = openpyxl.load_workbook(filename =self.filePathName)
        worksheet= wb.active

        worksheet["E"+ str(self.x)] = MonayWallet
        worksheet["F"+ str(self.x)] = Psum
        worksheet["G"+ str(self.x)] = Pl
        worksheet["H"+ str(self.x)] = Pg
        worksheet["I"+ str(self.x)] = Whsb
        worksheet["J"+ str(self.x)] = SOChsb



        for q in range (self.numberOfCar):
            colume= get_column_letter(12+2*q)


        self.x+=1
        wb = openpyxl.load_workbook(filename =self.filePathName)
        worksheet= wb.active
        
        colume= get_column_letter(14+3*numberOfCar)
        worksheet[str(colume)+str(self.x)] = Pcar[numberOfCar]
        worksheet[str(colume)+str(self.x)] = Pcar[numberOfCar+1]

        wb.save(self.filePathName)

