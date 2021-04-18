
from openpyxl import Workbook                    # openpyxl  2.6.2
from openpyxl.utils import get_column_letter

import datetime

basedate = datetime.datetime.today() 

# create some demo dates, roughly -5 to +5 months
some_dates = [basedate + datetime.timedelta(days = i*30) for i in range(-5,6)]
print(some_dates)

# create a workbook
wb = Workbook() 
ws1 = wb.active
ws1.title = "dates"
ws1["A1"] = "Daaaaaaaaates"
# fill dates manually to enable cell formatting
for i, date in enumerate(some_dates,2):
    ws1[f"A{i}"] = date                # no format

    ws1[f"B{i}"] = date                # formatted to MM.YY
    cell = ws1.cell(column=2, row=i)   # get cell and change format
    cell.number_format = "MM.YY"       # use 'MM-YY' if you want a dash between month/year

    # uncomment if you want to store the stringified version directly
    # ws1[f"C{i}"] = date.strftime("%m.%y")

wb.save(filename = 'workbook.xlsx')