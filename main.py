import time
import carBattery,homeStorageBattery,batteryManegmentSystem,linkEthNetwork,savingMeasurements,address
import numpy as np

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

UserGethUnlockAccount=0
Http='http://localhost:8545'
AddrEB=address.addressConcractElectricityBilling
AddrSC=address.addressConcractSystemRegulation
PathUserInfo='./ImportData/userInfo.xlsx'
PathUserSchedule='./ImportData/userSchedule.xlsx'
PathAbiSC='./SmartConcract/abiSystemControlingConcract.json'
PathAbiEB='./SmartConcract/abiElectricityBillingConcract.json'

t=1
dt=30
Day=1   
Hour=0
Min=0
Sec=0
Flg=1
StartFlg=True
k=1000

TarInt=0
SysRun=False
SysNedEne=False

EnergyMeter=0

ReqArrPower=[0]*5
SndArrPower=[0]*5
GetArrPower=[0]*5
ActArrPower=[0]*5

ActArrGrdPower=[0]*5
SumGrdPower=0
SumProPower=0
SumConPower=0

SumArrGrdEnergy=[0]*5
SumGrdEnergy=0
SumProEnergy=0
SumConEnergy=0


#Init moduls parameters
ethReg=linkEthNetwork.systemControling(AddrSC,PathAbiSC,Http,UserGethUnlockAccount)
ethBil=linkEthNetwork.electricityBilling(AddrEB,PathAbiEB,Http,UserGethUnlockAccount)

#Registration in SmartConcract

if ethReg.getUserIndex()==0:
    ethReg.autoRegistrationNewUser()
    time.sleep(5)

if ethBil.getUserIndex()==0:
    ethBil.autoRegistrationNewUser()
    time.sleep(5)


UserNumber=ethReg.getUserIndex()

#Init all parameters BMS
bms=batteryManegmentSystem.batteryManegmentSystem()

#Init all parameters Home storage Batery
Hsb=homeStorageBattery.homeStorageBattery(UserNumber,PathUserInfo)


#----------------------INIT PROPERTISE USER CARS------------------------------------------
wb = load_workbook(filename = PathUserInfo)
xlsUserInfo = wb['userCarProperties']
NumberOfCars=int(xlsUserInfo["C"+str(UserNumber+3)].value)

Car=[0]*NumberOfCars
for q in range (NumberOfCars):
    Car[q]=carBattery.carBattery(UserNumber,q,PathUserInfo)


#Init parameters for saving values
TimeOfTest = time.strftime("%H-%d-%m-%Y")
sm=savingMeasurements.savingMeasurements(UserNumber,TimeOfTest,NumberOfCars)

#---------------------------READ PARAMETERS FROM ETH NETWORK-----------------------------

SysRun=ethReg.getSystemRuning()
SysNedEne=ethReg.getSystemNeedsEnergy()

r=0
#----------------------OPEN FOLDER SCHEDULE USER---------------------------------
while r<123:

    StartTime=time.time_ns()

    wb = load_workbook(filename = PathUserSchedule)
    xlsxUserSchedule = wb["User "+str(UserNumber)]
    row=UserNumber+3

    row=((Day-1)*24)+Hour+4

#-------------------LOOKING DURATION PRICE ENERGY TARIFF-------------------------
    TarNum=xlsxUserSchedule["D"+str(row)].value
    TarInt=0
    for q in range (24):
        rowLop=row+q
        TarLop=xlsxUserSchedule["D"+str(rowLop)].value

        if (TarLop==TarNum):
            TarInt+=1
        else:
            break
#-----------------------POWER PROM DEVICE AND PV--------------------------------

    ReqPdSr=(xlsxUserSchedule["E"+str(row)].value)*1000
    ReqPdLd=(xlsxUserSchedule["F"+str(row)].value)*1000

    if ReqPdSr>ReqPdLd:
        HomNedEne=False
    else:
        HomNedEne=True


#-------------------LOOKING HOME AND CARS SETTINGS ------------------------------
    print ("")
    print("BATTERY SETINGS:")
    
    SOCsmart=xlsxUserSchedule["G"+str(row)].value
    Hsb.processBatterySetting(SOCsmart,Day,Hour,TarNum,TarInt,HomNedEne,SysNedEne)
    ReqPhsb=Hsb.getRequiredPower()

    ReqPcar=[0]*3

    for q in range (NumberOfCars):

        BatOn=0
        BatSet=0
        SOCstart=0

        try:
            colume= get_column_letter(8+3*q)
            BatOn=xlsxUserSchedule[str(colume)+str(row)].value

            colume= get_column_letter(9+3*q)
            BatSet=xlsxUserSchedule[str(colume)+str(row)].value

            colume= get_column_letter(10+3*q)
            SOCstart=xlsxUserSchedule[str(colume)+str(row)].value

        except:
            BatSet=0
            SOCstart=0

        Car[q].processingBatterySetting(BatOn,BatSet,SOCstart,Day,Hour,TarNum,HomNedEne,SysNedEne)
        ReqOnePcar=Car[q].getRequiredPower()
        ReqPcar=np.add(ReqPcar,ReqOnePcar)


#---------------------TOTAL CONSUPTION ----------------------
    
    ReqPbat=np.add(ReqPcar,ReqPhsb)

    ReqArrPower[0]=ReqPdSr
    ReqArrPower[1]=ReqPdLd
    ReqArrPower[2]=ReqPbat[0]
    ReqArrPower[3]=ReqPbat[1]
    ReqArrPower[4]=ReqPbat[2]

    print ("")
    print ("REQUAST POWERS:")
    print("PdSr:"+str(round(ReqArrPower[0]/k,2))+"kW  PdLd:"+str(round(ReqArrPower[1]/k,2))+"kW  PbAvSr:"
    +str(round(ReqArrPower[2]/k,2))+"kW  PbAvLd:"+str(round(ReqArrPower[3]/k,2))+"kW  PbRqLd:"
    +str(round(ReqArrPower[4]/k,2))+"kW")


#----------CHECK LIMITATIONS HAUSE MAX POWER WITH BMS---------------

    bms.processAllParametersAndRestrictions(ReqArrPower,GetArrPower)
    SndReqPower=bms.inputPowerDataInfoForConcract()

#------------SEND AND GET INFO POWER FROM ETH NETWORK----------------

    if ethReg.checkBlock():
        #Send demanded and requasted data:
        ethReg.setUserDataPower(SndReqPower)
        ethReg.modifaySystemTarifeNumber(TarNum)

        SysNedEne=ethReg.getSystemNeedsEnergy()

        #Get assagned data
        GetArrPower=ethReg.getUserDataPower()
    
    bms.processAllParametersAndRestrictions(ReqArrPower, GetArrPower)
        
#------------------GET ACTUAL POWERS-----------------------------------
    
    ActArrTotPower=bms.actualTotalPower()
    ActArrGrdPower=bms.actualPowerFromOrToGrid()

    ActGrdPower=0
    ActProPower=0
    ActConPower=0

    for q in range(5):

        if (q==0 or q==2):
            ActProPower+=ActArrTotPower[q]
        else:
            ActConPower+=ActArrTotPower[q]


    SumGrdPower+=ActConPower-ActProPower
    SumProPower+=ActProPower
    SumConPower+=ActConPower
    puActArrPower=bms.peerUnitRequestedPower()

    print ("")
    print ("ACTUAL POWERS IN SEGMENT:")
    print("PdSr:"+str(round(ActArrTotPower[0]/k,2))+"kW  PdLd:"+str(round(ActArrTotPower[1]/k,2))+"kW  PbAvSr:"
    +str(round(ActArrTotPower[2]/k,2))+"kW  PbAvLd:"+str(round(ActArrTotPower[3]/k,2))+"kW  PbRqLd:"
    +str(round(ActArrTotPower[4]/k,2))+"kW")
    print ("")
    print ("ACTUAL POWERS FLOW:")
    print("Ppro:"+str(round((ActProPower)/k,2))+"kW  Pcon:"+str(round(ActConPower/k,2))+"kW  Pgrd:"
    +str(round((ActConPower-ActProPower)/k,2))+"kW ")


#---------------- SET POWER INFO ON BATERY -----------------

    Hsb.setBatteryPower(puActArrPower)

    for q in range (NumberOfCars):
        Car[q].setBatteryPower(puActArrPower)

#--------------- TIME SLEEP-------------------
    Flg+=1
    Sec+=dt

    if Sec>=60:
        Min+=1
        Sec=0
    if Min>=60:
        Hour+=1
        Min=0
    if Hour>=24:
        Day+=1
        Hour=0
    if Day>7:
        Day=1

#----------------------UPDATE VALUES -----------------------

    SumArrGrdEnergy+=np.multiply(ActArrGrdPower,dt)
    
    EnergyMeter+=(ActConPower-ActProPower)*dt

    SumGrdEnergy+=((ActConPower-ActProPower)*dt)
    SumProEnergy+=(ActProPower*dt)
    SumConEnergy+=(ActConPower*dt)

    Hsb.updateBatteryValues(dt)

    for q in range (NumberOfCars):
        Car[q].updateBatteryValues(dt)

    ActGrdPower=0
    ActProPower=0
    ActConPower=0

#-------- SEND  ENERGY INFO IN ETEHREUM NETWORK ---------

    if ((Min==0 or Min==15 or Min==30 or Min==45) and Sec==dt):

        if Min==0 and StartFlg==False:
            row=((Day-1)*24)+Hour+3
            TarNumPre=xlsxUserSchedule["D"+str(row)].value
        else:
            row=((Day-1)*24)+Hour+4
            TarNumPre=xlsxUserSchedule["D"+str(row)].value

        wb.close()

        StartFlg=False

        wb = load_workbook(filename = PathUserInfo)
        xlsxSystemTarifPrices = wb["systemTariffPrices"]
        PriceBuy=xlsxSystemTarifPrices["C"+str(TarNumPre+2)].value
        PriceSell=xlsxSystemTarifPrices["D"+str(TarNumPre+2)].value
        wb.close()

        ethBil.modifaySystemTarifPrice(int(TarNumPre),int(PriceBuy), int(PriceSell))
        ethBil.setUserDataEnergy([int(SumArrGrdEnergy[0]),int(SumArrGrdEnergy[1]),int(SumArrGrdEnergy[2]),int(SumArrGrdEnergy[3]),int(SumArrGrdEnergy[4])])
        
        SumArrGrdEnergy=[0]*5

#------------------Safe measurment of energy and avg power-------------------------

        AvgGrdPower=(SumGrdPower/Flg)
        AvgProPower=(SumProPower/Flg)
        AvgConPower=(SumConPower/Flg)

        sm.safeBasicMeasurements(Day,Hour,Min,SumGrdEnergy,SumProEnergy,SumConEnergy,AvgGrdPower,AvgProPower,AvgConPower)

        if (Hsb.BatOn==True):
            InfoBat=Hsb.getBatteryInfo(Flg)
            sm.safeHomeBatteryMeasurements(InfoBat)

        if (NumberOfCars>0):

            for q in range (NumberOfCars):
                InfoBat=Car[q].getBatteryInfo(Flg)
                sm.safeCarMeasurements(q, InfoBat)


        SumGrdEnergy=0
        SumConEnergy=0
        SumProEnergy=0

        SumGrdPower=0
        SumProPower=0
        SumConPower=0

        Flg=0
        
#---------------------ETH BILING PREVIOUS SENDED PRICE-------------------------

    if ((Min==5 or Min==20 or Min==35 or Min==50) and Sec==dt):

        ethBil.processingBillingForEnergy()

#----------------------SAVE PRICE AND WALLET WALLEUS----------------------------------

    if ((Min==10 or Min==25 or Min==40 or Min==55) and Sec==dt):

        MonayWalletCent=ethBil.getUserWalletInCent()
        PriceForEnergyCent=ethBil.getUserFinalEnergyPriceInCent()

        sm.safeCashBalance(MonayWalletCent, PriceForEnergyCent)

#-------------------SAVE MEASURMENT ALSO IN EXE FILE---------------
    print("")
    print("TIME")
    print(str(Day)+":"+str(Hour)+":"+str(Min)+":"+str(Sec))
    print("")
    print("---------------------------------------------------------")

    while (StartTime+t*1000000000)>time.time_ns():
        None