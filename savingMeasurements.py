import openpyxl
from openpyxl import Workbook, worksheet, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill, Alignment

class savingMeasurements(object):
    def __init__(self, UserNumber,TestNumber,NumberOfCars):
        self.NumberOfCars=NumberOfCars
        self.x=3
        self.FilePathName='./ExportData/Data User'+str(UserNumber)+' Test'+str(TestNumber)+'.xlsx'

        wb = openpyxl.Workbook()

        PowerWorksheet = wb.create_sheet("PowerMeausurments")
        EnergyWorksheet= wb.create_sheet("EnergyMeausurments")

        self.fontStyleWord = Font(name="Calibri",size = "7")
        self.fontStyleNumber=Font(name="Calibri",size = "10")

        self.alignmentStyle=Alignment(horizontal='center',vertical='center')


        EnergyWorksheet.cell(row = self.x, column = 2, value = 'Time').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 3, value = 'Wallet[€]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 4, value = 'Price[€]').font = self.fontStyleWord

        EnergyWorksheet.cell(row = self.x, column = 5, value = 'Ein[kWh]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 6, value = 'Eout[kWh]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 7, value = 'EdSr[kWh]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 8, value = 'EdLd[kWh]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 9, value = 'EbAvSr[kWh]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 10, value = 'EbAvLd[kWh]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 11, value = 'EbRqLd[kWh]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 12, value = 'Ehsb[kW/h]').font = self.fontStyleWord
        EnergyWorksheet.cell(row = self.x, column = 13, value = 'SOChsb[%]').font = self.fontStyleWord


        PowerWorksheet.cell(row = self.x, column = 2, value = 'Time').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 3, value = 'Wallet[€]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 4, value = 'Price[€]').font = self.fontStyleWord

        PowerWorksheet.cell(row = self.x, column = 5, value = 'Pin[kW]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 6, value = 'Pout[kW]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 7, value = 'PdSr[kW]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 8, value = 'PdLd[kW]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 9, value = 'PdAvSr[kW]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 10, value = 'PdAvLd[kW]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 11, value = 'PdRqLd[kW]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 12, value = 'Phsb[kW]').font = self.fontStyleWord
        PowerWorksheet.cell(row = self.x, column = 13, value = 'SOChsb[%]').font = self.fontStyleWord

        for q in range (self.NumberOfCars):
            EnergyWorksheet.cell(row = self.x, column = 14+3*q, value = 'Ecar'+str(self.NumberOfCars)+'[kWh]').font = self.fontStyleWord
            EnergyWorksheet.cell(row = self.x, column = 15+3*q, value = 'SOCcar'+str(self.NumberOfCars)+'[%]').font = self.fontStyleWord

            PowerWorksheet.cell(row = self.x, column = 14+3*q, value = 'Pcar'+str(self.NumberOfCars)+'[kW]').font = self.fontStyleWord
            PowerWorksheet.cell(row = self.x, column = 15+3*q, value = 'SOCcar'+str(self.NumberOfCars)+'[%]').font = self.fontStyleWord
            numberOfCell=10+3*q


        for q in range (2,16+3*q):
            EnergyWorksheet.cell(row = self.x, column = q).alignment=self.alignmentStyle
            PowerWorksheet.cell(row = self.x, column = q).alignment=self.alignmentStyle
            colume= get_column_letter(q)
            EnergyWorksheet.column_dimensions[colume].width =9.4
            PowerWorksheet.column_dimensions[colume].width =9.4

            EnergyWorksheet.column_dimensions["B"].width =14
            PowerWorksheet.column_dimensions["B"].width =14



        wb.save(filename = self.FilePathName)
        wb.close()

        wb = openpyxl.load_workbook(filename =self.FilePathName)
        Sheetworksheet= wb["Sheet"]
        wb.remove(Sheetworksheet)
        wb.save(filename = self.FilePathName)
        wb.close()

    def safeBasicMeasurements(self,DataTime,AvgPin,AvgPout,AvgArrTotPower,SumEin,SumEout,SumArrTotEnergy):

        self.x+=1
        wb = openpyxl.load_workbook(filename =self.FilePathName)
        PowerWorksheet = wb["PowerMeausurments"]
        EnergyWorksheet= wb["EnergyMeausurments"]

        k=1000 #convert W to kW
        h=3600 #convert Ws to Wh

        EnergyWorksheet.cell(row = self.x, column = 2, value = DataTime).font = self.fontStyleNumber
        PowerWorksheet.cell(row = self.x, column = 2, value = DataTime).font= self.fontStyleNumber


        EnergyWorksheet.cell(row = self.x, column = 5, value = "%.3f" % round(SumEin/(k*h),3)).font = self.fontStyleNumber
        EnergyWorksheet.cell(row = self.x, column = 6, value = "%.3f" % round(SumEout/(k*h),3)).font = self.fontStyleNumber
        EnergyWorksheet.cell(row = self.x, column = 7, value = "%.3f" % round(SumArrTotEnergy[0]/(k*h),3)).font = self.fontStyleNumber
        EnergyWorksheet.cell(row = self.x, column = 8, value = "%.3f" % round(SumArrTotEnergy[1]/(k*h),3)).font = self.fontStyleNumber
        EnergyWorksheet.cell(row = self.x, column = 9, value = "%.3f" % round(SumArrTotEnergy[2]/(k*h),3)).font = self.fontStyleNumber
        EnergyWorksheet.cell(row = self.x, column = 10, value = "%.3f" % round(SumArrTotEnergy[3]/(k*h),3)).font = self.fontStyleNumber
        EnergyWorksheet.cell(row = self.x, column = 11, value = "%.2f" % round(SumArrTotEnergy[4]/(k*h),3)).font = self.fontStyleNumber

        PowerWorksheet.cell(row = self.x, column = 5, value = "%.3f" % round(AvgPin/k,3)).font = self.fontStyleNumber
        PowerWorksheet.cell(row = self.x, column = 6, value = "%.3f" % round(AvgPout/k,3)).font = self.fontStyleNumber
        PowerWorksheet.cell(row = self.x, column = 7, value = "%.3f" % round(AvgArrTotPower[0]/k,3)).font = self.fontStyleNumber
        PowerWorksheet.cell(row = self.x, column = 8, value = "%.3f" % round(AvgArrTotPower[1]/k,3)).font = self.fontStyleNumber
        PowerWorksheet.cell(row = self.x, column = 9, value = "%.3f" % round(AvgArrTotPower[2]/k,3)).font = self.fontStyleNumber
        PowerWorksheet.cell(row = self.x, column = 10, value = "%.3f" % round(AvgArrTotPower[3]/k,3)).font = self.fontStyleNumber
        PowerWorksheet.cell(row = self.x, column = 11, value = "%.3f" % round(AvgArrTotPower[4]/k,3)).font = self.fontStyleNumber


        wb.save(filename = self.FilePathName)
        wb.close()



    def safeHomeBatteryMeasurements(self,InfoBat):

        k=1000 #convert W to kW
        h=3600 #convert Ws to Wh
        p=100
        wb = openpyxl.load_workbook(filename =self.FilePathName)
        PowerWorksheet = wb["PowerMeausurments"]
        EnergyWorksheet= wb["EnergyMeausurments"]

        PowerWorksheet.cell(row = self.x, column = 12, value = "%.3f" % round(InfoBat[0]/k,3)).font = self.fontStyleNumber
        PowerWorksheet.cell(row = self.x, column = 13, value = "%.3f" % round(InfoBat[2]*p,2)).font = self.fontStyleNumber

        EnergyWorksheet.cell(row = self.x, column = 12, value = "%.2f" % round(InfoBat[1]/(k*h),3)).font = self.fontStyleNumber
        EnergyWorksheet.cell(row = self.x, column = 13, value = "%.2f" % round(InfoBat[2]*p,2)).font = self.fontStyleNumber


        wb.save(filename = self.FilePathName)
        wb.close()

    def safeCarBatteryMeasurements(self,CarNumber,InfoBat):

        k=1000 #convert W to kW
        h=3600 #convert Ws to Wh
        p=100
        wb = openpyxl.load_workbook(filename =self.FilePathName)
        PowerWorksheet = wb["PowerMeausurments"]
        EnergyWorksheet= wb["EnergyMeausurments"]

        PowerWorksheet.cell(row = self.x, column = 14+(3*CarNumber), value = "%.3f" % round(InfoBat[0]/k,3)).font = self.fontStyleNumber
        PowerWorksheet.cell(row = self.x, column = 15+(3*CarNumber), value = "%.2f" % round(InfoBat[2]*p,2)).font = self.fontStyleNumber

        EnergyWorksheet.cell(row = self.x, column = 14+(3*CarNumber), value = "%.3f" % round(InfoBat[1]/(k*h),3)).font = self.fontStyleNumber
        EnergyWorksheet.cell(row = self.x, column = 15+(3*CarNumber), value = "%.2f" % round(InfoBat[2]*p,2)).font = self.fontStyleNumber

        wb.save(filename = self.FilePathName)
        wb.close()


    def safeCashBalance(self,MonayWalletCent, PriceForEnergyCent):

        wb = openpyxl.load_workbook(filename =self.FilePathName)
        PowerWorksheet = wb["PowerMeausurments"]
        EnergyWorksheet= wb["EnergyMeausurments"]

        EnergyWorksheet.cell(row = self.x, column = 3, value = "%.2f" % round(MonayWalletCent/100,3)).font = self.fontStyleNumber
        EnergyWorksheet.cell(row = self.x, column = 4, value = "%.2f" % round(PriceForEnergyCent/100,3)).font = self.fontStyleNumber

        PowerWorksheet.cell(row = self.x, column = 3, value = "%.2f" % round(MonayWalletCent/100,3)).font = self.fontStyleNumber
        PowerWorksheet.cell(row = self.x, column = 4, value = "%.2f" % round(PriceForEnergyCent/100,3)).font = self.fontStyleNumber


        wb.save(filename = self.FilePathName)
        wb.close()


    def deletePreviousValues(self):

        wb = openpyxl.load_workbook(filename =self.FilePathName)
        worksheet= wb["SystemMeausurments"]

        for q in (3, self.x + 1):
            for i in range(2, 16+3*self.CarNumber):
                worksheet.delete_rows(q, i)

        wb.save(filename = self.FilePathName)
        wb.close()     